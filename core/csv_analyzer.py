"""
Structured spreadsheet analysis for CSV and Excel uploads.
"""

import csv
import io
import statistics
from collections import Counter


REQUIRED_REPORT_HEADINGS = (
    "Executive Summary",
    "Important Statistics",
    "Category Breakdown",
    "Key Insights",
    "Recommendations",
)


def _try_float(value: str) -> float | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None

    normalized = text.replace(",", "")
    try:
        return float(normalized)
    except ValueError:
        return None


def _detect_dialect(sample: str):
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        class _FallbackDialect:
            delimiter = ","
            quotechar = '"'
            doublequote = True
            skipinitialspace = True
            lineterminator = "\r\n"
            quoting = csv.QUOTE_MINIMAL

        return _FallbackDialect()


def _humanize(name: str) -> str:
    cleaned = str(name or "").replace("_", " ").replace("-", " ").strip()
    return cleaned[:1].upper() + cleaned[1:] if cleaned else "Unnamed Column"


def _read_csv_rows(raw: str) -> list[list[str]]:
    if not raw or not raw.strip():
        return []

    dialect = _detect_dialect(raw[:8192])
    try:
        rows = list(csv.reader(io.StringIO(raw), dialect=dialect))
    except Exception:
        rows = list(csv.reader(io.StringIO(raw)))

    return [[str(cell or "").strip() for cell in row] for row in rows if any(str(cell or "").strip() for cell in row)]


def _normalize_rows(rows: list[list[str]]) -> tuple[list[str], list[list[str]]]:
    if len(rows) < 2:
        return [], []

    max_columns = max(len(row) for row in rows)
    padded_rows = [row + [""] * (max_columns - len(row)) for row in rows]
    header = [cell or f"column_{index + 1}" for index, cell in enumerate(padded_rows[0])]
    body = padded_rows[1:]
    return header, body


def is_data_table(raw: str, threshold: float = 0.55) -> bool:
    rows = _read_csv_rows(raw)
    header, body = _normalize_rows(rows)
    if len(header) < 2 or len(body) < 2:
        return False

    structured_columns = 0
    for column_index in range(len(header)):
        values = [row[column_index].strip() for row in body if row[column_index].strip()]
        if not values:
            structured_columns += 1
            continue

        numeric_count = sum(1 for value in values if _try_float(value) is not None)
        average_length = sum(len(value) for value in values) / len(values)
        unique_count = len(set(values))
        is_categorical = average_length < 40 and unique_count <= max(25, len(values) * 0.5)

        if numeric_count / len(values) > 0.5 or is_categorical:
            structured_columns += 1

    return (structured_columns / max(len(header), 1)) >= threshold


def _analyze_numeric(values: list[float]) -> dict[str, float]:
    count = len(values)
    if count == 0:
        return {}

    mean = statistics.fmean(values)
    median = statistics.median(values)
    minimum = min(values)
    maximum = max(values)
    stdev = statistics.pstdev(values) if count > 1 else 0.0
    return {
        "count": count,
        "mean": round(mean, 2),
        "median": round(median, 2),
        "min": round(minimum, 2),
        "max": round(maximum, 2),
        "stdev": round(stdev, 2),
    }


def _analyze_categorical(values: list[str], top_k: int = 3) -> dict[str, object]:
    counter = Counter(values)
    total = sum(counter.values())
    top_values = [
        (label, count, round((count / total) * 100, 1))
        for label, count in counter.most_common(top_k)
    ]
    return {
        "count": total,
        "unique": len(counter),
        "top": top_values,
    }


def _correlation(xs: list[float], ys: list[float]) -> float | None:
    if len(xs) < 3 or len(xs) != len(ys):
        return None

    mean_x = statistics.fmean(xs)
    mean_y = statistics.fmean(ys)
    numerator = sum((x - mean_x) * (y - mean_y) for x, y in zip(xs, ys))
    denominator_x = sum((x - mean_x) ** 2 for x in xs) ** 0.5
    denominator_y = sum((y - mean_y) ** 2 for y in ys) ** 0.5
    if denominator_x == 0 or denominator_y == 0:
        return None
    return numerator / (denominator_x * denominator_y)


def _collect_column_values(header: list[str], body: list[list[str]]):
    columns = {name: [] for name in header}
    for row in body:
        for index, name in enumerate(header):
            columns[name].append((row[index] if index < len(row) else "").strip())
    return columns


def _build_dataset_report(header: list[str], body: list[list[str]], source_label: str) -> str:
    if not header or not body:
        return "\n".join([
            "Executive Summary",
            "- No spreadsheet data was detected.",
            "",
            "Important Statistics",
            "- No numeric statistics are available.",
            "",
            "Category Breakdown",
            "- No categorical columns were detected.",
            "",
            "Key Insights",
            "- The dataset needs at least one header row and one data row.",
            "",
            "Recommendations",
            "- Add more rows or columns and try again.",
        ])

    columns = _collect_column_values(header, body)
    numeric_columns: dict[str, list[float]] = {}
    categorical_columns: dict[str, list[str]] = {}

    for name, values in columns.items():
        non_empty = [value for value in values if value]
        if not non_empty:
            continue

        numeric_values = [number for number in (_try_float(value) for value in non_empty) if number is not None]
        if numeric_values and len(numeric_values) / len(non_empty) >= 0.5:
            numeric_columns[name] = numeric_values
            continue

        average_length = sum(len(value) for value in non_empty) / len(non_empty)
        unique_count = len(set(non_empty))
        if average_length < 40 and unique_count <= max(25, len(non_empty) * 0.6):
            categorical_columns[name] = non_empty

    strongest_correlation = None
    numeric_names = list(numeric_columns.keys())
    for left_index in range(len(numeric_names)):
        for right_index in range(left_index + 1, len(numeric_names)):
            left_name = numeric_names[left_index]
            right_name = numeric_names[right_index]
            left_column = header.index(left_name)
            right_column = header.index(right_name)
            pairs = []
            for row in body:
                left_value = _try_float(row[left_column]) if left_column < len(row) else None
                right_value = _try_float(row[right_column]) if right_column < len(row) else None
                if left_value is not None and right_value is not None:
                    pairs.append((left_value, right_value))

            if len(pairs) < 3:
                continue

            xs = [pair[0] for pair in pairs]
            ys = [pair[1] for pair in pairs]
            correlation = _correlation(xs, ys)
            if correlation is None:
                continue

            if strongest_correlation is None or abs(correlation) > abs(strongest_correlation[2]):
                strongest_correlation = (left_name, right_name, correlation)

    report: list[str] = []

    report.append("Executive Summary")
    report.append(
        f"- This {source_label} contains {len(body)} records across {len(header)} columns."
    )
    report.append(
        f"- Fields analyzed: {', '.join(_humanize(name) for name in header[:8])}"
        + ("." if len(header) <= 8 else ", and more.")
    )
    if numeric_columns:
        report.append(
            f"- Numeric columns detected: {', '.join(_humanize(name) for name in numeric_columns.keys())}."
        )
    else:
        report.append("- No numeric columns were detected.")

    report.append("")
    report.append("Important Statistics")
    if numeric_columns:
        for name, values in numeric_columns.items():
            stats = _analyze_numeric(values)
            report.append(
                f"- {_humanize(name)}: average {stats['mean']}, median {stats['median']}, "
                f"min {stats['min']}, max {stats['max']}, standard deviation {stats['stdev']}."
            )
    else:
        report.append("- No numeric statistics are available for this spreadsheet.")

    report.append("")
    report.append("Category Breakdown")
    if categorical_columns:
        for name, values in categorical_columns.items():
            analysis = _analyze_categorical(values)
            top_values = ", ".join(
                f"{label} ({percentage}%)"
                for label, _count, percentage in analysis["top"]
            )
            report.append(
                f"- {_humanize(name)}: {analysis['unique']} distinct values; most common entries are {top_values}."
            )
    else:
        report.append("- No categorical columns were detected.")

    report.append("")
    report.append("Key Insights")
    if strongest_correlation is not None:
        left_name, right_name, correlation = strongest_correlation
        direction = "positive" if correlation > 0 else "negative"
        report.append(
            f"- {_humanize(left_name)} and {_humanize(right_name)} show the strongest {direction} relationship ({correlation:+.2f})."
        )
    if numeric_columns:
        widest_range_name = max(
            numeric_columns,
            key=lambda name: max(numeric_columns[name]) - min(numeric_columns[name]),
        )
        widest_range = max(numeric_columns[widest_range_name]) - min(numeric_columns[widest_range_name])
        report.append(
            f"- {_humanize(widest_range_name)} has the widest numeric spread at {round(widest_range, 2)}."
        )
    if categorical_columns:
        largest_categorical_name = max(categorical_columns, key=lambda name: len(set(categorical_columns[name])))
        report.append(
            f"- {_humanize(largest_categorical_name)} has the most category variety with "
            f"{len(set(categorical_columns[largest_categorical_name]))} distinct values."
        )
    if len(report) > 0 and report[-1] == "Key Insights":
        report.append("- The spreadsheet is mostly text and should be reviewed column by column.")

    report.append("")
    report.append("Recommendations")
    if numeric_columns:
        report.append("- Review outliers in the numeric columns with the largest ranges.")
    else:
        report.append("- Add more measurable values if you need deeper statistical analysis.")
    if categorical_columns:
        report.append("- Monitor the dominant categories for imbalance or unexpected concentration.")
    else:
        report.append("- Add grouped labels or status fields if you need category analysis.")
    report.append("- Validate missing values and header names before sharing this analysis.")

    return "\n".join(report)


def analyze_csv(raw: str) -> str:
    rows = _read_csv_rows(raw)
    header, body = _normalize_rows(rows)
    return _build_dataset_report(header, body, source_label="CSV dataset")


def analyze_workbook_bytes(data: bytes) -> str:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(data), data_only=True)
    selected_sheet_name = None
    selected_rows: list[list[str]] = []

    for sheet in workbook.worksheets:
        rows = []
        for row in sheet.iter_rows(values_only=True):
            values = ["" if value is None else str(value).strip() for value in row]
            if any(values):
                rows.append(values)

        if len(rows) >= 2:
            selected_sheet_name = sheet.title
            selected_rows = rows
            break

    if not selected_rows:
        return _build_dataset_report([], [], source_label="Excel workbook")

    header, body = _normalize_rows(selected_rows)
    return _build_dataset_report(
        header,
        body,
        source_label=f"Excel sheet '{selected_sheet_name}'",
    )


def is_structured_dataset_report(text: str) -> bool:
    normalized = (text or "").strip()
    return all(heading in normalized for heading in REQUIRED_REPORT_HEADINGS)
